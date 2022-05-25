from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from inspector import Code


class ExcelWriter:
    def __init__(self, sample_id):
        self.file_path = '/Users/suyeon/Downloads/offset_inspect_result_{}.xlsx'.format(sample_id)
        self.sheet_error = None
        self.excel = Workbook()

    def get_offset_inspect_result(self, error_groups):
        self.create_sheet()
        self.change_column_width()
        self.write_column()
        for group in error_groups:
            self.write_group_data(group)
            self.write_each_data(group)
        self.save()

    def create_sheet(self):
        sheet = self.excel.active
        sheet.title = Code.SHEET_ERROR
        self.sheet_error = self.excel[Code.SHEET_ERROR]

    def change_column_width(self):
        for column in range(1, 6):
            self.sheet_error.column_dimensions[get_column_letter(column)].width = 25

    def write_column(self):
        self.sheet_error.append(Code.EXCEL_COLUMN_SPEECH)
        column_cell_list = ['A1', 'B1', 'C1', 'D1', 'E1']
        for cell in column_cell_list:
            self.fill_grey_for_title_cells(cell)

    def write_group_data(self, group):
        group_data = [Code.SENTENCE, group.sentence_value, group.sentence_offset_start, group.sentence_offset_end,
                      group.sentence_value_with_offset]
        self.excel[Code.SHEET_ERROR].append(group_data)

        row = self.excel[Code.SHEET_ERROR].max_row
        self.bold_font_for_content_cells(row)

    def write_each_data(self, group):
        for word in group.error_words:
            each_data = [word[Code.T_VALUE_INDEX], word[Code.S_VALUE_INDEX], word[Code.OFFSET_START_INDEX],
                         word[Code.OFFSET_END_INDEX]]
            self.excel[Code.SHEET_ERROR].append(each_data)

    def fill_grey_for_title_cells(self, cell):
        grey = PatternFill(fill_type='solid', start_color='00C0C0C0', end_color='00C0C0C0')
        self.sheet_error[cell].fill = grey
        self.sheet_error[cell].alignment = Alignment(horizontal="center", vertical="center")
        self.sheet_error[cell].font = Font(bold=True)

    def bold_font_for_content_cells(self, row):
        for column in range(1, 6):
            self.sheet_error.cell(row=row, column=column).font = Font(bold=True)
            self.sheet_error.cell(row=row, column=column).border = Border(top=Side(border_style="thick", color="000000"))

    def save(self):
        self.excel.save(self.file_path)
        print("검사 결과를 엑셀로 저장하였습니다.")
